#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
HRE 연결마스터 Excel 파일 구조 생성 스크립트
생성 일시: 2026-01-21
용도: VBA 임포트 전 Excel 파일의 시트, 테이블, 범위 구조 자동 생성
"""

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_table(ws, table_name, ref, headers, style="TableStyleMedium2"):
    """Excel ListObject 테이블 생성"""
    # 헤더 작성
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 테이블 객체 생성
    tab = Table(displayName=table_name, ref=ref)
    style_info = TableStyleInfo(
        name=style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style_info
    ws.add_table(tab)

    return ws

def main():
    print("🔧 HRE 연결마스터 Excel 파일 생성 시작...")

    # 1. Workbook 생성
    wb = Workbook()
    wb.remove(wb.active)  # 기본 Sheet 제거

    # 2. 13개 시트 생성 (순서대로)
    sheets_config = [
        ("Guide", True),
        ("CoAMaster", True),
        ("CorpMaster", True),
        ("CorpCoA", True),
        ("BSPL", True),
        ("ADBS", True),
        ("Verify", True),
        ("Check", True),
        ("HideSheet", False),  # 숨김
        ("DirectoryURL", False),  # 숨김
        ("Memo", False),  # 숨김
        ("AddCoA", True),  # 나중에 VBA가 숨김
        ("AddCoA_ADBS", True),  # 나중에 VBA가 숨김
    ]

    for sheet_name, visible in sheets_config:
        ws = wb.create_sheet(title=sheet_name)
        if not visible:
            ws.sheet_state = 'hidden'
        print(f"  ✓ 시트 생성: {sheet_name} {'(숨김)' if not visible else ''}")

    # 3. CoAMaster - Master 테이블
    print("\n📋 테이블 생성 중...")
    ws_coa = wb["CoAMaster"]
    create_table(
        ws_coa,
        "Master",
        "A1:K1",
        ["Account", "Description", "연결계정명", "분류", "Category", "BSPL", "대분류", "Ranking", "부호", "금액", "비고"]
    )
    # 컬럼 너비 조정
    ws_coa.column_dimensions['A'].width = 10  # Account
    ws_coa.column_dimensions['B'].width = 35  # Description
    ws_coa.column_dimensions['C'].width = 20  # 연결계정명
    ws_coa.column_dimensions['D'].width = 15  # 분류
    ws_coa.column_dimensions['E'].width = 20  # Category
    ws_coa.column_dimensions['F'].width = 8   # BSPL
    ws_coa.column_dimensions['G'].width = 10  # 대분류
    ws_coa.column_dimensions['H'].width = 10  # Ranking
    ws_coa.column_dimensions['I'].width = 8   # 부호
    ws_coa.column_dimensions['J'].width = 15  # 금액
    ws_coa.column_dimensions['K'].width = 20  # 비고
    print("  ✓ CoAMaster.Master (11 cols)")

    # 4. CorpMaster - Corp 테이블
    ws_corp = wb["CorpMaster"]
    create_table(
        ws_corp,
        "Corp",
        "A1:J1",
        ["법인코드", "법인명", "Entity Name", "Hierarchy", "Scope", "취득일", "처분일", "지분율", "기능통화", "Consolidation Method"]
    )
    ws_corp.column_dimensions['A'].width = 12  # 법인코드
    ws_corp.column_dimensions['B'].width = 25  # 법인명
    ws_corp.column_dimensions['C'].width = 30  # Entity Name
    ws_corp.column_dimensions['D'].width = 12  # Hierarchy
    ws_corp.column_dimensions['E'].width = 8   # Scope
    ws_corp.column_dimensions['F'].width = 12  # 취득일
    ws_corp.column_dimensions['G'].width = 12  # 처분일
    ws_corp.column_dimensions['H'].width = 10  # 지분율
    ws_corp.column_dimensions['I'].width = 12  # 기능통화
    ws_corp.column_dimensions['J'].width = 20  # Consolidation Method
    print("  ✓ CorpMaster.Corp (10 cols)")

    # 5. CorpCoA - Raw_CoA 테이블
    ws_corpcoa = wb["CorpCoA"]
    create_table(
        ws_corpcoa,
        "Raw_CoA",
        "A1:I1",
        ["법인코드", "계정코드", "연결계정명", "Reporting COA", "Account", "Description", "Variant Type", "Internal Transaction Flag", "비고"]
    )
    ws_corpcoa.column_dimensions['A'].width = 12  # 법인코드
    ws_corpcoa.column_dimensions['B'].width = 20  # 계정코드
    ws_corpcoa.column_dimensions['C'].width = 20  # 연결계정명
    ws_corpcoa.column_dimensions['D'].width = 15  # Reporting COA
    ws_corpcoa.column_dimensions['E'].width = 10  # Account
    ws_corpcoa.column_dimensions['F'].width = 30  # Description
    ws_corpcoa.column_dimensions['G'].width = 15  # Variant Type
    ws_corpcoa.column_dimensions['H'].width = 20  # Internal Transaction Flag
    ws_corpcoa.column_dimensions['I'].width = 20  # 비고
    print("  ✓ CorpCoA.Raw_CoA (9 cols)")

    # 6. BSPL - PTB 테이블
    ws_bspl = wb["BSPL"]
    create_table(
        ws_bspl,
        "PTB",
        "A1:H1",
        ["법인코드", "계정코드", "계정과목명", "차변", "대변", "차변-대변", "PwC_CoA", "PwC_계정과목명"]
    )
    ws_bspl.column_dimensions['A'].width = 12  # 법인코드
    ws_bspl.column_dimensions['B'].width = 15  # 계정코드
    ws_bspl.column_dimensions['C'].width = 25  # 계정과목명
    ws_bspl.column_dimensions['D'].width = 15  # 차변
    ws_bspl.column_dimensions['E'].width = 15  # 대변
    ws_bspl.column_dimensions['F'].width = 15  # 차변-대변
    ws_bspl.column_dimensions['G'].width = 12  # PwC_CoA
    ws_bspl.column_dimensions['H'].width = 25  # PwC_계정과목명
    print("  ✓ BSPL.PTB (8 cols)")

    # 7. ADBS - AD_BS 테이블
    ws_adbs = wb["ADBS"]
    create_table(
        ws_adbs,
        "AD_BS",
        "A1:I1",
        ["법인코드", "계정코드", "계정과목명", "취득일자", "처분일자", "차변", "대변", "PwC_CoA", "PwC_계정과목명"]
    )
    ws_adbs.column_dimensions['A'].width = 12  # 법인코드
    ws_adbs.column_dimensions['B'].width = 15  # 계정코드
    ws_adbs.column_dimensions['C'].width = 25  # 계정과목명
    ws_adbs.column_dimensions['D'].width = 12  # 취득일자
    ws_adbs.column_dimensions['E'].width = 12  # 처분일자
    ws_adbs.column_dimensions['F'].width = 15  # 차변
    ws_adbs.column_dimensions['G'].width = 15  # 대변
    ws_adbs.column_dimensions['H'].width = 12  # PwC_CoA
    ws_adbs.column_dimensions['I'].width = 25  # PwC_계정과목명
    print("  ✓ ADBS.AD_BS (9 cols)")

    # 8. HideSheet - 4개 테이블
    ws_hide = wb["HideSheet"]

    # 테이블 1: 결산연월 (A1:B2)
    ws_hide['A1'] = "결산연도"
    ws_hide['B1'] = "결산월"
    ws_hide['A2'] = 2026
    ws_hide['B2'] = 1
    tab1 = Table(displayName="결산연월", ref="A1:B2")
    style_info = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab1.tableStyleInfo = style_info
    ws_hide.add_table(tab1)
    ws_hide.column_dimensions['A'].width = 12
    ws_hide.column_dimensions['B'].width = 10

    # 테이블 2: Link (D1:E2)
    ws_hide['D1'] = "SPO_Link"
    ws_hide['E1'] = "Path"
    ws_hide['D2'] = "https://pwckor.sharepoint.com/sites/KR-ASR-HRE_Consolidation"
    ws_hide['E2'] = ""
    tab2 = Table(displayName="Link", ref="D1:E2")
    tab2.tableStyleInfo = style_info
    ws_hide.add_table(tab2)
    ws_hide.column_dimensions['D'].width = 60
    ws_hide.column_dimensions['E'].width = 30

    # 테이블 3: 비경상적 (G1:H1)
    create_table(ws_hide, "비경상적", "G1:H1", ["PwC_CoA", "PwC_계정과목명"])
    ws_hide.column_dimensions['G'].width = 12
    ws_hide.column_dimensions['H'].width = 25

    # 테이블 4: 환율마스터 (J1:K1)
    create_table(ws_hide, "환율마스터", "J1:K1", ["통화", "환율"])
    ws_hide.column_dimensions['J'].width = 10
    ws_hide.column_dimensions['K'].width = 15

    # N2 셀: 버전 정보 위치 (VBA가 자동으로 채움)
    ws_hide['N1'] = "AppVersion"
    ws_hide['N1'].font = Font(bold=True)
    ws_hide.column_dimensions['N'].width = 15

    print("  ✓ HideSheet.결산연월 (2 cols)")
    print("  ✓ HideSheet.Link (2 cols)")
    print("  ✓ HideSheet.비경상적 (2 cols)")
    print("  ✓ HideSheet.환율마스터 (2 cols)")

    # 9. AddCoA - CoA_Input 테이블
    ws_addcoa = wb["AddCoA"]
    create_table(
        ws_addcoa,
        "CoA_Input",
        "A1:G1",
        ["법인코드", "법인별CoA", "법인별계정과목명", "PwC_CoA", "PwC_계정과목명", "적요", "비고"]
    )
    ws_addcoa.column_dimensions['A'].width = 12
    ws_addcoa.column_dimensions['B'].width = 15
    ws_addcoa.column_dimensions['C'].width = 25
    ws_addcoa.column_dimensions['D'].width = 12
    ws_addcoa.column_dimensions['E'].width = 25
    ws_addcoa.column_dimensions['F'].width = 20
    ws_addcoa.column_dimensions['G'].width = 20
    print("  ✓ AddCoA.CoA_Input (7 cols)")

    # 10. AddCoA_ADBS - CoA_Input_ADBS 테이블
    ws_addcoa_adbs = wb["AddCoA_ADBS"]
    create_table(
        ws_addcoa_adbs,
        "CoA_Input_ADBS",
        "A1:G1",
        ["법인코드", "법인별CoA", "법인별계정과목명", "PwC_CoA", "PwC_계정과목명", "적요", "비고"]
    )
    ws_addcoa_adbs.column_dimensions['A'].width = 12
    ws_addcoa_adbs.column_dimensions['B'].width = 15
    ws_addcoa_adbs.column_dimensions['C'].width = 25
    ws_addcoa_adbs.column_dimensions['D'].width = 12
    ws_addcoa_adbs.column_dimensions['E'].width = 25
    ws_addcoa_adbs.column_dimensions['F'].width = 20
    ws_addcoa_adbs.column_dimensions['G'].width = 20
    print("  ✓ AddCoA_ADBS.CoA_Input_ADBS (7 cols)")

    # 11. Check 시트 - 워크플로 상태 추적
    print("\n📋 Check 시트 워크플로 생성 중...")
    ws_check = wb["Check"]

    # 헤더 행
    headers = ["단계", "작업명", "설명", "상태", "작업일시", "작업자"]
    for idx, header in enumerate(headers, start=1):
        cell = ws_check.cell(row=1, column=idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 12개 워크플로 단계
    workflow_steps = [
        (1, "SPO 설정", "SharePoint URL 설정"),
        (2, "조직 설정", "부서 코드 설정"),
        (3, "결산연월 설정", "2026년 1월 설정"),
        (4, "법인 추가", "대상 법인 등록"),
        (5, "연결범위 설정", "Scope 지정"),
        (6, "CoA 마스터 검토", "178개 계정 확인"),
        (7, "CoA 확인 및 데이터 합산", "PTB 데이터 로드"),
        (8, "CoA 추가/수정/삭제", "매핑 완료"),
        (9, "환율 조회", "평균/기말 환율 조회"),
        (10, "합산 검증", "차변=대변 검증"),
        (11, "취득/처분 CoA 확인", "ADBS 데이터 로드"),
        (12, "취득/처분 검증", "ADBS 검증"),
    ]

    for row_idx, (step_num, task_name, description) in enumerate(workflow_steps, start=12):
        ws_check.cell(row=row_idx, column=1, value=step_num)
        ws_check.cell(row=row_idx, column=2, value=task_name)
        ws_check.cell(row=row_idx, column=3, value=description)
        ws_check.cell(row=row_idx, column=4, value="")  # 상태 빈칸
        ws_check.cell(row=row_idx, column=5, value="")  # 작업일시 빈칸
        ws_check.cell(row=row_idx, column=6, value="")  # 작업자 빈칸

    # 컬럼 너비 조정
    ws_check.column_dimensions['A'].width = 8   # 단계
    ws_check.column_dimensions['B'].width = 25  # 작업명
    ws_check.column_dimensions['C'].width = 30  # 설명
    ws_check.column_dimensions['D'].width = 15  # 상태
    ws_check.column_dimensions['E'].width = 20  # 작업일시
    ws_check.column_dimensions['F'].width = 15  # 작업자

    # 테두리 추가
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws_check['A1:F23']:
        for cell in row:
            cell.border = thin_border

    print("  ✓ Check 시트 워크플로 (12 단계)")

    # 12. Guide 시트 - 안내 문서
    ws_guide = wb["Guide"]
    ws_guide['A1'] = "HRE 연결마스터 사용 가이드"
    ws_guide['A1'].font = Font(size=16, bold=True, color="4472C4")
    ws_guide['A3'] = "이 파일은 HRE 그룹의 연결재무제표 작성을 위한 Excel 기반 시스템입니다."
    ws_guide['A5'] = "시작하기:"
    ws_guide['A6'] = "1. Alt+F11을 눌러 VBA 편집기를 엽니다."
    ws_guide['A7'] = "2. 파일 → 가져오기를 통해 VBA_Export 폴더의 모든 .bas 파일을 임포트합니다."
    ws_guide['A8'] = "3. UserForms 폴더의 모든 .frm 파일을 임포트합니다."
    ws_guide['A9'] = "4. Custom UI Editor로 리본 메뉴 XML을 추가합니다."
    ws_guide['A10'] = "5. 상세한 내용은 '완벽한_구현_가이드.md' 파일을 참조하세요."
    ws_guide['A12'] = "문의: https://github.com/busyway1/HRE.git"
    ws_guide.column_dimensions['A'].width = 80
    print("  ✓ Guide 시트 안내 문서 추가")

    # 13. 빈 시트 (Verify, DirectoryURL, Memo)
    print("\n📋 빈 시트 설정 완료: Verify, DirectoryURL, Memo")

    # 14. 파일 저장
    output_path = "/Users/jaewookim/Desktop/Project/HRE/작업/연결마스터_HRE_v1.00.xlsm"
    wb.save(output_path)
    print(f"\n✅ 파일 생성 완료: {output_path}")
    print(f"   - 총 13개 시트 생성")
    print(f"   - 총 11개 ListObject 테이블 생성")
    print(f"   - Check 시트 워크플로 12단계 설정")
    print(f"   - HideSheet, DirectoryURL, Memo 숨김 처리")
    print(f"\n🎯 다음 단계: VBA 모듈 임포트 (Alt+F11 → 파일 → 가져오기)")

    return output_path

if __name__ == "__main__":
    main()
